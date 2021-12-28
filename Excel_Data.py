# Read a Excel file which cointeins information about bad eggs

import openpyxl    
# we creat a dictionary like that: {"IMG_982933_09876.jpg":[6,2,0,0,0,0,0,0,0,0],...}

def read_excel(path): #// Abrimos  el excel segun la ruta especificada 
    wb_obj = openpyxl.load_workbook(path)#//Abrimos el excel
    sheet_obj = wb_obj.active
    m_column = sheet_obj.max_column #// Obtenemos las columnas maximas del excel
    m_row=sheet_obj.max_row #// Obtenemos las filas máximas del excel
    dic={}#// Inicamos un diccionario vacio 
    for x in range(2,m_row+1): #// Recorremos las filas empezando por la segunda. La primera fila contiene el nombre
        images=[]#// Iniciamos una lista 
        for y in range(1, m_column + 1):#// Recorremos las columnas empezando por la primera.
            cell_obj = sheet_obj.cell(row = x, column =y)#// Obtenemos el valor de cada celda.
            images.append(cell_obj.value)#// Guardamos el nombre de la foto y las posiciones de los huevos malos en forma de lista// ['IMG_20210309_105623.jpg', 6, 2, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1] 
        dic[images[0]]=images[1:]#// Convertimos la lista anterior en diccionario con clave el nombre de la imagen y como valor las posiciones de los huevos malos // {'IMG_20210309_095911.jpg': [6, 2, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0]}
    return dic #// Devolvemos un diccionario con todas las imagenes y las posiciones de los huevos malos

def bad_egg (name_egg,dic): #Enter a dicctionary with bad eggs
    if name_egg in dic: #// Comprobamos que la foto contenga huevos malos
        lista=dic[name_egg][2:]#// Guardamos la lista de huevos malos [0,1,0,0,0,1.....]
        pos=[i+1 for i, e in enumerate(lista) if e == 1]#// Apartir de la lista [0,1,0,0,0,1...] obtenemos la posicion [2,6]
        return pos #// Devolevmos la posicion en forma de lista.
    else:
        return False #// Si la imagen no contiene huevos malos devolvemos un False

if __name__=='__main__':
    dic=read_excel(r'C:\Users\Erick\Desktop\TFG\Documentos\analisi_rafa.xlsx')